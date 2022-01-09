#importing the Employee.xlsx using openpyxl
from openpyxl import workbook, load_workbook
wb = load_workbook("Employeedata.xlsx")
ws = wb.active
for i in range(2,ws.max_row + 1):
    cell = ws.cell(i, 2)
    if "helpinghands.cm" in cell.value:
        #replacing th helpinghands.cm with handsinhands.org
        updated_email = (cell.value).replace("helpinghands.cm", "handsinhands.org")
        #replacing the handsinhands.org in the second cell, that is 
        #the cell where our helpinghands.cm is found
        ws.cell(i, 2).value = updated_email
#new.xls is the output file created
    wb.save("new.xls")    